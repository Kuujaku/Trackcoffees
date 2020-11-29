import openpyxl
from os import path
from datetime import date


def load_workbook(fln):
    if path.exists(fln):
        return openpyxl.load_workbook(fln)
    return openpyxl.Workbook()


fln = "kahvit.xlsx"  # edit filename to your liking
workbook = load_workbook(fln)
sheet = workbook.active

sheet["A1"] = "Kauppa"
sheet["B1"] = "Nimi"
sheet["C1"] = "Hinta"
sheet["D1"] = "Määrä"
sheet["E1"] = "Jauhatus"
sheet["F1"] = "Tyyppi"
sheet["G1"] = "Paahto"
sheet["H1"] = "Makuprofiili"
sheet["I1"] = "Arvostelu"
sheet["J1"] = "Arvosana"
sheet["K1"] = "Alkuperämaa"
sheet["L1"] = "Päivämäärä"
sheet["M1"] = "Hinnat yhteensä: "


usin1 = input("Kaupan nimi: ")
usin2 = input("Tuotteen nimi: ")
usin3 = int(input("Tuotteen hinta: "))  # use only numbers if you want the sum
usin4 = input("Tuotteen määrä: ")
usin5 = input("Kahvin jauhatus: ")
usin6 = input("Arabica vai Robusta (esim 50/50): ")
usin7 = input("Kahvin paahtoaste: ")
usin8 = input("Kahvin maku: ")
usin9 = input("Lyhyt arvostelu: ")
usin10 = input("Arvosana: ")
usin11 = input("Tuotteen alkuperämaa: ")


today = date.today()
parsed_date = today.strftime("%d/%m/%Y")
paiva = parsed_date

cell = sheet.cell(row=1, column=14)
cell.value = "=SUM(C2:C100)"  # counts the money spend sum


usin = [
    (usin1, usin2, usin3, usin4, usin5, usin6,
     usin7, usin8, usin9, usin10, usin11, paiva)
]

for user in usin:
    sheet.append(user)


while True:
    ifprint = input("Do you want to print all entries? (Y)es (N)o: ")
    if ifprint in ['Y', 'N']:
        break

if ifprint == 'Y':
    for row in sheet.iter_cols(min_row=1, max_col=12):
        for cell in row:
            print(cell.value, end=" ")
        print()

if ifprint == 'N':
    pass

workbook.save(fln)
